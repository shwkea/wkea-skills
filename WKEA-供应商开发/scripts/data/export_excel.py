import os
import sys
import json
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

def export_excel(workspace_path):
    """导出供应商核查清单Excel"""
    
    # 读取suppliers.json
    suppliers_json_path = os.path.join(workspace_path, 'suppliers.json')
    if not os.path.exists(suppliers_json_path):
        print(json.dumps({"status": "error", "message": "suppliers.json不存在"}, ensure_ascii=False))
        return
    
    with open(suppliers_json_path, 'r', encoding='utf-8') as f:
        suppliers_data = json.load(f)
    
    # 读取product.json
    product_json_path = os.path.join(workspace_path, '产品信息', 'product.json')
    product_name = ""
    if os.path.exists(product_json_path):
        with open(product_json_path, 'r', encoding='utf-8') as f:
            product_data = json.load(f)
            product_name = product_data.get("name", "")
    
    # 创建工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "供应商核查清单"
    
    # 标题行
    ws['A1'] = f"供应商开发核查清单 - {product_name}"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:G1')
    
    # 表头
    headers = ['序号', '供应商名称', '来源渠道', '是否授权', '价格', '货期', '核查状态']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')
    
    # 数据行
    for idx, supplier in enumerate(suppliers_data.get("suppliers", []), 1):
        ws.cell(row=3+idx, column=1, value=idx)
        ws.cell(row=3+idx, column=2, value=supplier.get("name", ""))
        ws.cell(row=3+idx, column=3, value=supplier.get("source", ""))
        ws.cell(row=3+idx, column=4, value="是" if supplier.get("is_authorized") else "否")
        ws.cell(row=3+idx, column=5, value=supplier.get("price") or "待确认")
        ws.cell(row=3+idx, column=6, value=supplier.get("delivery") or "待确认")
        
        # 核查状态
        verification = supplier.get("verification", {})
        status = "已完成" if all(v == "verified" for v in verification.values()) else "进行中"
        ws.cell(row=3+idx, column=7, value=status)
    
    # 调整列宽
    column_widths = [8, 30, 15, 12, 15, 15, 12]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64+col)].width = width
    
    # 保存
    output_path = os.path.join(workspace_path, '供应商开发核查清单_汇总.xlsx')
    wb.save(output_path)
    
    print(json.dumps({"status": "success", "path": output_path}, ensure_ascii=False))

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("用法: python export_excel.py <工作目录>")
        sys.exit(1)
    
    workspace_path = sys.argv[1]
    export_excel(workspace_path)
